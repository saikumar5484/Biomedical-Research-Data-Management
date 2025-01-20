import openpyxl
from openpyxl import Workbook
import random
import os

# Function to generate fake project details
def generate_project_details():
    types = ["Research", "Development", "Documentation", "Study", "Evaluation"]
    topics = ["Cardiology", "Neurology", "Oncology", "Pediatrics", "Surgery"]
    return f"{random.choice(types)} on {random.choice(topics)}"

# Function to distribute projects across years
def distribute_projects(start_year, end_year, total_projects):
    years = list(range(start_year, end_year + 1))
    year_project_counts = {year: 0 for year in years}

    for _ in range(total_projects):
        year = random.choice(years)
        year_project_counts[year] += 1

    return year_project_counts

# Start year, end year, and total projects
start_year = 1987
end_year = 2024
total_projects = 1200  # Total projects to distribute

# Base document path
base_document_path = "path/to/fake/documents"

# Distribute projects across years
year_project_counts = distribute_projects(start_year, end_year, total_projects)

# Create Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Hospital Projects Data"

# Add headers
headers = ["S.No", "Year", "Document Link", "Project ID", "Project Details"]
ws.append(headers)

# Add project data
serial_number = 1
for year, project_count in year_project_counts.items():
    for _ in range(project_count):
        document_link = f"{base_document_path}/project_{serial_number}.pdf"
        project_id = f"PID-{serial_number:04d}"  # Unique project ID
        project_details = generate_project_details()
        ws.append([serial_number, year, document_link, project_id, project_details])
        serial_number += 1

# Dynamically get the location of manage.py
current_directory = os.path.dirname(os.path.abspath(__file__))  # Directory where the script is located
file_path = os.path.join(current_directory, "Hospital_Projects_Data1.xlsx")

# Debug: Verify workbook data
print(f"Sheet names: {wb.sheetnames}")
print(f"First row of data: {list(ws.iter_rows(min_row=1, max_row=1, values_only=True))}")
print(f"Attempting to save file to: {file_path}")

# Save the workbook with error handling
try:
    wb.save(file_path)
    print(f"File saved successfully at: {file_path}")
except Exception as e:
    print(f"Failed to save file: {e}")

# Suggest a fallback directory if saving fails
fallback_path = os.path.expanduser("~/Desktop/Hospital_Projects_Data1.xlsx")
if not os.path.exists(file_path):
    try:
        wb.save(fallback_path)
        print(f"File saved to fallback location: {fallback_path}")
    except Exception as e:
        print(f"Failed to save file to fallback location: {e}")
